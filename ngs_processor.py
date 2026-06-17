"""
NGS下机数据自动化处理脚本
功能：质控检查、SNVIndel过滤统计、生成报告Excel
"""

import os
import sys
import glob
import time
import warnings
from copy import copy
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 颜色常量
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF9999", end_color="FF9999")
YELLOW_FILL = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
ORANGE_FILL = PatternFill("solid", start_color="FFCC99", end_color="FFCC99")
HEADER_FILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
BLUE_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
GREY_FILL   = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")

HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
BOLD_FONT    = Font(bold=True, name="Arial", size=10)
RED_FONT     = Font(bold=True, color="FF0000", name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


# ─────────────────────────────────────────────
# QC 阈值
# ─────────────────────────────────────────────
QC_RULES = {
    "CleanQ30":           (">=", 0.75,  ""),
    "Depth_CDS":          (">=", 400.0, ""),
    "RNA-Control":        (">=", 20.0,  ""),
    "Coverage(50x)_SNP":  (">=", 0.90,  ""),
}

QC_REPORT_COLS = [
    "Sample", "CleanData", "CleanQ30", "Depth_CDS",
    "Coverage(50x)_SNP", "RNA-Control",
    "MSI_Ratio", "MSI_Num", "MSI_State",
    "ContaRatio", "ContaStat",
]

SNV_REVIEW_COLS = [
    "Sample", "Chr", "Start", "End", "Ref", "Alt",
    "Tags", "Depth", "Freq", "AltDepth",
    "Gene", "Type", "CDSChange", "Amplicon", "Plus", "Minus",
]

PASSTHROUGH_SHEETS = ["HD_pass", "Fusion"]   # AmpliconStat handled separately as pivot

# ─────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────
def to_num(val):
    """安全转换为浮点数，去除 % 符号"""
    try:
        return float(str(val).replace("%", "").strip())
    except Exception:
        return None

def bytes_to_human(val):
    """
    将数字（如 23000000000）转换为带单位字符串（如 "23.0G"）。
    若原始值已含字母单位（如 "2.3G"）则原样返回。
    单位：T / G / M / K，保留1位小数，去掉尾零。
    """
    s = str(val).strip()
    if s and s[-1].upper() in ("T", "G", "M", "K", "B"):
        return s
    try:
        n = float(s.replace(",", ""))
    except Exception:
        return s
    for unit, threshold in [("T", 1e12), ("G", 1e9), ("M", 1e6), ("K", 1e3)]:
        if abs(n) >= threshold:
            formatted = f"{n / threshold:.1f}".rstrip("0").rstrip(".")
            return f"{formatted}{unit}"
    return s

def style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row_idx, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        if fill:
            cell.fill = fill
        cell.font      = NORMAL_FONT
        cell.alignment = LEFT
        cell.border    = THIN_BORDER


def auto_col_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def freeze_header(ws):
    ws.freeze_panes = ws["A2"]


# ─────────────────────────────────────────────
# QC 检查
# ─────────────────────────────────────────────
def check_qc(summary_df: pd.DataFrame):
    """
    返回:
      qc_df        — 仅保留 QC_REPORT_COLS 的 DataFrame
      fail_dict    — {sample: [不合格项, ...]}
      pass_samples — set
    """
    available_cols = [c for c in QC_REPORT_COLS if c in summary_df.columns]
    qc_df = summary_df[available_cols].copy()

    fail_dict = {}
    for _, row in summary_df.iterrows():
        sample = str(row.get("Sample", "Unknown"))
        fails  = []
        for col, (op, threshold, unit) in QC_RULES.items():
            if col not in row.index:
                continue
            val = to_num(row[col])
            if val is None:
                fails.append(f"{col}(数据缺失)")
                continue
            if op == ">=" and val < threshold:
                fails.append(f"{col}={val}{unit}(质控标准≥{threshold}{unit})")
        if fails:
            fail_dict[sample] = fails

    pass_samples = set(summary_df["Sample"].astype(str)) - set(fail_dict.keys())
    return qc_df, fail_dict, pass_samples


# ─────────────────────────────────────────────
# SNVIndel 处理
# ─────────────────────────────────────────────
def process_snvindel(snv_df: pd.DataFrame, discard_df: pd.DataFrame, fake_pos_df: pd.DataFrame | None) -> pd.DataFrame:
    """
    1. 过滤 Tags 含 Black_list / Polymorphism
    2. 统计三列：DisCard出现次数 / SNVIndel内CDSChange次数 / 是否假阳
    3. 标记 AltDepth <= 30
    返回处理后 DataFrame（含新增列）
    """
    # 1. 过滤黑名单
    if "Tags" in snv_df.columns:
        mask = snv_df["Tags"].astype(str).str.contains(
            r"Black_list|Polymorphism", case=False, na=False
        )
        snv_df = snv_df[~mask].copy()

    # 构建唯一识别键
    def make_key(df):
        gene = df["Gene"].astype(str).str.strip() if "Gene" in df.columns else pd.Series([""] * len(df))
        cds  = df["CDSChange"].astype(str).str.strip() if "CDSChange" in df.columns else pd.Series([""] * len(df))
        return gene + "|" + cds

    snv_df["_key"] = make_key(snv_df)

    # 2a. DisCard 出现次数（Gene+CDSChange）
    discard_count = {}
    if discard_df is not None and not discard_df.empty:
        discard_df["_key"] = make_key(discard_df)
        discard_count = discard_df["_key"].value_counts().to_dict()
    snv_df["DisCard_Count"] = snv_df["_key"].map(lambda k: discard_count.get(k, 0))

    # 2b. SNVIndel 内 CDSChange 出现次数（全表，Gene+CDSChange键）
    inner_count = snv_df["_key"].value_counts().to_dict()
    snv_df["SNVIndel_Count"] = snv_df["_key"].map(lambda k: inner_count.get(k, 0))

    # 2c. 是否在假阳文件中
    fake_keys = set()
    if fake_pos_df is not None and not fake_pos_df.empty:
        fake_pos_df["_key"] = make_key(fake_pos_df)
        fake_keys = set(fake_pos_df["_key"])
    snv_df["IsFakePositive"] = snv_df["_key"].map(lambda k: "是" if k in fake_keys else "否")

    # 3. AltDepth 标记
    if "AltDepth" in snv_df.columns:
        snv_df["LowAltDepth_Flag"] = snv_df["AltDepth"].apply(
            lambda x: "★低AltDepth" if (pd.notna(x) and to_num(x) is not None and to_num(x) <= 30) else ""
        )

    snv_df.drop(columns=["_key"], inplace=True, errors="ignore")
    return snv_df


# ─────────────────────────────────────────────
# 写 QC_Report Sheet
# ─────────────────────────────────────────────
def write_qc_report(wb, qc_df, fail_dict, summary_df):
    ws = wb.create_sheet("QC_Report")
    ws.sheet_view.showGridLines = True

    headers = list(qc_df.columns)
    ncols   = len(headers)

    # 写表头
    ws.append(headers)
    style_header_row(ws, 1, ncols)

    # QC阈值列索引（用于条件格式）
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # CleanData 列索引（保留 "2.3G" 字符串格式，强制文本对齐）
    cleandata_col = col_idx.get("CleanData")

    for r_idx, (_, row) in enumerate(qc_df.iterrows(), start=2):
        sample = str(row.get("Sample", ""))
        is_fail = sample in fail_dict
        row_fill = RED_FILL if is_fail else None

        # 构建行数据，CleanData 转为 "2.3G" 格式字符串
        row_vals = []
        for c in headers:
            if c == "CleanData":
                row_vals.append(bytes_to_human(row.get(c, "")))
            else:
                row_vals.append(row.get(c, ""))
        ws.append(row_vals)
        style_data_row(ws, r_idx, ncols, fill=row_fill)

        # CleanData 单元格强制文本格式，防止 Excel 将字符串解析回数字
        if cleandata_col:
            cell = ws.cell(row=r_idx, column=cleandata_col)
            cell.number_format = "@"
            cell.alignment = LEFT

    # 条件格式：数值列 < 阈值 → 红底
    thresholds = {
        "CleanQ30":          0.75,
        "Depth_CDS":         400.0,
        "RNA-Control":       20.0,
        "Coverage(50x)_SNP": 0.90,
    }
    nrows = len(qc_df) + 1
    for col_name, threshold in thresholds.items():
        if col_name not in col_idx:
            continue
        c = col_idx[col_name]
        col_letter = get_column_letter(c)
        cell_range = f"{col_letter}2:{col_letter}{nrows}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=[str(threshold)], fill=RED_FILL)
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThanOrEqual", formula=[str(threshold)], fill=GREEN_FILL)
        )

    freeze_header(ws)
    auto_col_width(ws)
    ws.row_dimensions[1].height = 20


# ─────────────────────────────────────────────
# 写 QC_FailItems Sheet
# ─────────────────────────────────────────────
def write_qc_failitem(wb, fail_dict):
    ws = wb.create_sheet("QC_FailItems")
    ws.append(["Sample", "QC_FailMessage"])
    style_header_row(ws, 1, 2)

    if not fail_dict:
        ws.append(["—", "所有样本均通过质控"])
        style_data_row(ws, 2, 2, fill=GREEN_FILL)
    else:
        for r_idx, (sample, fails) in enumerate(fail_dict.items(), start=2):
            msg = f"{sample} {', '.join(fails)} 质控不合格"
            ws.append([sample, msg])
            style_data_row(ws, r_idx, 2, fill=RED_FILL)

    freeze_header(ws)
    auto_col_width(ws)


# ─────────────────────────────────────────────
# 写 SNVIndel_Review Sheet
# ─────────────────────────────────────────────
def write_snvindel_review(wb, snv_df):
    ws = wb.create_sheet("SNVIndel_Review")

    # 确定输出列顺序：原始保留列 + 新增三列 + 标记列
    extra_cols = ["DisCard_Count", "SNVIndel_Count", "IsFakePositive", "LowAltDepth_Flag"]
    available_review = [c for c in SNV_REVIEW_COLS if c in snv_df.columns]
    available_extra  = [c for c in extra_cols if c in snv_df.columns]
    all_cols = available_review + available_extra

    out_df = snv_df[all_cols].copy()
    ncols  = len(all_cols)
    headers = list(all_cols)

    ws.append(headers)
    style_header_row(ws, 1, ncols)

    # 列索引
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    for r_idx, (_, row) in enumerate(out_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])

        low_flag = str(row.get("LowAltDepth_Flag", ""))
        fake_pos = str(row.get("IsFakePositive", ""))
        discard  = to_num(row.get("DisCard_Count", 0)) or 0

        if low_flag:
            fill = ORANGE_FILL
        elif fake_pos == "是":
            fill = YELLOW_FILL
        elif discard > 0:
            fill = BLUE_FILL
        else:
            fill = None

        style_data_row(ws, r_idx, ncols, fill=fill)

    # 条件格式：AltDepth <= 30 整行橙色
    if "AltDepth" in col_idx:
        c = col_idx["AltDepth"]
        col_letter = get_column_letter(c)
        nrows = len(out_df) + 1
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{nrows}",
            FormulaRule(
                formula=[f"${col_letter}2<=30"],
                fill=ORANGE_FILL,
            )
        )

    # DisCard_Count > 0 → 蓝色字体
    if "DisCard_Count" in col_idx:
        c = col_idx["DisCard_Count"]
        col_letter = get_column_letter(c)
        nrows = len(out_df) + 1
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{nrows}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=Font(bold=True, color="0070C0", name="Arial", size=10))
        )

    freeze_header(ws)
    auto_col_width(ws)
    ws.row_dimensions[1].height = 20


# ─────────────────────────────────────────────
# 写 SNVIndel_Full Sheet（保留过滤后全部列）
# ─────────────────────────────────────────────
def write_snvindel_full(wb, snv_df):
    ws = wb.create_sheet("SNVIndel_Full")
    ncols = len(snv_df.columns)
    headers = list(snv_df.columns)

    ws.append(headers)
    style_header_row(ws, 1, ncols)

    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    nrows_data = len(snv_df)

    for r_idx, (_, row) in enumerate(snv_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        style_data_row(ws, r_idx, ncols)

    # 条件格式：AltDepth <= 30 → 橙色
    if "AltDepth" in col_idx:
        c = col_idx["AltDepth"]
        col_letter = get_column_letter(c)
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{nrows_data + 1}",
            FormulaRule(formula=[f"${col_letter}2<=30"], fill=ORANGE_FILL)
        )

    freeze_header(ws)
    auto_col_width(ws)

# ─────────────────────────────────────────────
# CNV Sheet：复制后标记 Confidence=Low 行
# ─────────────────────────────────────────────
def write_cnv(wb, cnv_df: pd.DataFrame):
    if cnv_df is None or cnv_df.empty:
        return
    ws      = wb.create_sheet("CNV")
    headers = list(cnv_df.columns)
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    ws.append(headers)
    style_header_row(ws, 1, ncols)

    for r_idx, (_, row) in enumerate(cnv_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        is_low = str(row.get("Confidence", "")).strip().lower() == "low"
        style_data_row(ws, r_idx, ncols, fill=ORANGE_FILL if is_low else None)

    # 动态条件格式
    if "Confidence" in col_idx:
        cl = get_column_letter(col_idx["Confidence"])
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{len(cnv_df) + 1}",
            FormulaRule(formula=[f'${cl}2="Low"'], fill=ORANGE_FILL),
        )
    freeze_header(ws)
    auto_col_width(ws)


# ─────────────────────────────────────────────
# AmpliconStat Sheet：生成数据透视表
# 列=Sample，行=Amplicon，值=RoT
# ─────────────────────────────────────────────
def write_amplicon_pivot(wb, amplicon_df: pd.DataFrame):
    if amplicon_df is None or amplicon_df.empty:
        return

    required = {"Sample", "Amplicon", "RoT"}
    missing = required - set(amplicon_df.columns)
    ws = wb.create_sheet("AmpliconStat")
    if missing:
        print(f"  ⚠ AmpliconStat 缺少列 {missing}，按原始数据输出")
        headers = list(amplicon_df.columns)
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        for r_idx, (_, row) in enumerate(amplicon_df.iterrows(), start=2):
            ws.append([row.get(c, "") for c in headers])
            style_data_row(ws, r_idx, len(headers))
        freeze_header(ws)
        auto_col_width(ws)
        return

    # 数值转换
    df = amplicon_df.copy()
    df["RoT"] = pd.to_numeric(df["RoT"], errors="coerce")
    pivot = df.pivot_table(index="Amplicon", columns="Sample",
                           values="RoT", aggfunc="mean").reset_index()
    pivot = pivot.reset_index()
    pivot.columns.name = None

    samples = [c for c in pivot.columns if c != "Amplicon"]
    headers = ["Amplicon"] + samples
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    ws.append(headers)
    style_header_row(ws, 1, ncols)

    for r_idx, (_, row) in enumerate(pivot.iterrows(), start=2):
        ws.append([row.get(h, "") for h in headers])
        style_data_row(ws, r_idx, ncols)
        for s in samples:
            val = row.get(s)
            c   = col_idx[s]
            cell = ws.cell(row=r_idx, column=c)
            if pd.notna(val):
                cell.value        = round(float(val), 4)
                cell.number_format = "0.0000"
                cell.alignment    = CENTER
        ws.cell(row=r_idx, column=col_idx["Amplicon"]).alignment = LEFT

    ws.freeze_panes = ws["B2"]
    auto_col_width(ws)

# ═══════════════════════════════════════════════════════════════
# HD_pass / Fusion：直接复制 DataFrame（原始数据输出）
# ═══════════════════════════════════════════════════════════════

def write_passthrough(wb, sheet_name: str, df: pd.DataFrame):
    if df is None or df.empty:
        return
    ws      = wb.create_sheet(sheet_name)
    headers = list(df.columns)
    ncols   = len(headers)
    ws.append(headers)
    style_header_row(ws, 1, ncols)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        style_data_row(ws, r_idx, ncols)
    freeze_header(ws)
    auto_col_width(ws)


# ─────────────────────────────────────────────
# 主处理流程
# ─────────────────────────────────────────────
def process_file(ngs_path, fake_pos_df):
    t0 = time.time()
    print(f"\n{'='*60}")
    print(f"处理文件: {os.path.basename(ngs_path)}")
    print(f"{'='*60}")

    # 耗时最多的读入阶段
    t_read = time.time()
    summary_df = pd.read_excel(ngs_path, sheet_name="Summary")
    snv_df = pd.read_excel(ngs_path, sheet_name="SNVIndel")
    discard_df = pd.read_excel(ngs_path, sheet_name="SNVIndelDiscard")
    cnv_df = pd.read_excel(ngs_path, sheet_name="CNV")
    amplicon_df = pd.read_excel(ngs_path, sheet_name="AmpliconStat")
    print(f"  文件读取耗时: {time.time() - t_read:.1f}s")

    # ── Summary / QC ──
    qc_df, fail_dict = pd.DataFrame(), {}
    if summary_df.empty:
        print("  ⚠ 未找到 Summary Sheet，跳过质控")
    else:
        for col in QC_RULES:
            if col in summary_df.columns:
                summary_df[col] = summary_df[col].apply(
                    lambda x: to_num(x) if to_num(x) is not None else x
                )
        qc_df, fail_dict, pass_samples = check_qc(summary_df)
        print(f"  样本数: {len(summary_df)}, 质控通过: {len(pass_samples)}, 不合格: {len(fail_dict)}")

    # ── SNVIndel ──
    if snv_df.empty:
        print("  ⚠ 未找到 SNVIndel Sheet")
    else:
        snv_df = process_snvindel(snv_df, discard_df, fake_pos_df)
        print(f"  SNVIndel 过滤后行数: {len(snv_df)}")

    # ── 生成输出文件 ──
    t_write = time.time()
    base_name = os.path.splitext(os.path.basename(ngs_path))[0]
    out_path  = os.path.join(os.path.dirname(ngs_path), f"{base_name}_Report.xlsx")

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active) # 删除默认空白sheet

    # 写各Sheet
    if not qc_df.empty:
        write_qc_report(out_wb, qc_df, fail_dict, summary_df)
    write_qc_failitem(out_wb, fail_dict)

    if not snv_df.empty:
        write_snvindel_review(out_wb, snv_df)
        #write_snvindel_full(out_wb, snv_df)

    write_cnv(out_wb, cnv_df)

    raw_data = {}
    for sheet in PASSTHROUGH_SHEETS:
        raw_data[sheet] = pd.read_excel(ngs_path, sheet_name=sheet)
        if not raw_data[sheet].empty:
            write_passthrough(out_wb, sheet, raw_data[sheet])
            print(f"  已输出 Sheet: {sheet}")

    write_amplicon_pivot(out_wb, amplicon_df)

    out_wb.save(out_path)
    print(f"  写出耗时: {time.time() - t_write:.1f}s")

    elapsed = time.time() - t0
    print(f"\n  ✔ 报告已生成: {os.path.basename(out_path)}  （总耗时 {elapsed:.1f}s）")

    # 打印质控不合格汇总
    #if fail_dict:
    #    print("\n  【质控不合格样本】")
    #    for sample, fails in fail_dict.items():
    #        print(f"    {sample} {', '.join(fails)} 质控不合格")

def process_fake_pos_path(work_dir):
    fake_candidates = glob.glob(os.path.join(work_dir, "**", "CP200_db.xlsx"), recursive=True)
    fake_pos_path = fake_candidates[0] if fake_candidates else None
    if fake_pos_path and os.path.exists(fake_pos_path):
        print(f"找到假阳文件: {os.path.basename(fake_pos_path)}")
        try:
            fp_sheets = pd.read_excel(fake_pos_path, sheet_name=None, dtype=str)
            fp_first  = list(fp_sheets.values())[0]
            fake_pos_df = fp_first
            print(f"  已加载假阳文件: {os.path.basename(fake_pos_path)}")
            return fake_pos_df
        except Exception as e:
            print(f"  ⚠ 假阳文件读取失败: {e}")
    else:
        print("⚠ 未找到 CP200假阳SNV.xlsx，假阳性检查将跳过")
# ─────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────
def main():
    # 确定工作目录
    if getattr(sys, "frozen", False):
        work_dir = os.path.dirname(sys.executable)
    else:
        work_dir = os.getcwd()

    print("NGS下机数据自动化处理工具")
    print(f"工作目录: {work_dir}")

    # 查找假阳文件
    fake_pos_df = process_fake_pos_path(work_dir)

    # 查找 NGS 下机数据
    ngs_files = glob.glob(os.path.join(work_dir, "**", "*ADXHS-OncoPro*.xlsx"), recursive=True)
    # 排除已生成的报告文件
    ngs_files = [f for f in ngs_files if "_Report" not in os.path.basename(f)]

    if not ngs_files:
        print("\n⚠ 未找到包含 'ADXHS-OncoPro' 的xlsx文件")
        print("请将下机数据文件放在同一目录下")
        input("\n按回车键退出...")
        return

    print(f"\n共找到 {len(ngs_files)} 个下机数据文件")
    
    total_t0 = time.time()
    file_times = []

    for ngs_path in ngs_files:
        t0 = time.time()
        try:
            process_file(ngs_path, fake_pos_df)
        except Exception as e:
            print(f"\n  ✗ 处理文件 {os.path.basename(ngs_path)} 时出错: {e}")
            import traceback
            traceback.print_exc()
        file_times.append((os.path.basename(ngs_path), time.time() - t0))

    total_elapsed = time.time() - total_t0
    if len(ngs_files) > 1:
        print("各文件耗时：")
        for fname, t in file_times:
            print(f"  {fname}: {t:.1f} 秒")
    print(f"全部处理完成！总耗时 {total_elapsed:.1f} 秒")
    input("按回车键退出...")


if __name__ == "__main__":
    main()
