"""
Excel 报告写出模块
所有 write_* 函数通过参数接收数据和配置，不引用全局状态。
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from config import ProductConfig
from filters import to_num, QC_PASS, QC_RISK, QC_FAIL, _eval_qc_col

# ─────────────────────────────────────────────
# 样式常量
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF9999", end_color="FF9999")
YELLOW_FILL = PatternFill("solid", start_color="FFFF99", end_color="FFFF99")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
ORANGE_FILL = PatternFill("solid", start_color="FFCC99", end_color="FFCC99")
HEADER_FILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
BLUE_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
NORMAL_FONT = Font(name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


# ─────────────────────────────────────────────
# 通用样式工具
# ─────────────────────────────────────────────
def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER;  cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 20


def _style_row(ws, row_idx: int, ncols: int, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        if fill:
            cell.fill = fill
        cell.font = NORMAL_FONT; cell.alignment = LEFT; cell.border = THIN_BORDER


def _auto_col_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[cl].width = min(max(max_len + 2, min_w), max_w)


def _bytes_to_human(val) -> str:
    s = str(val).strip()
    if s and s[-1].upper() in ("T", "G", "M", "K", "B"):
        return s
    try:
        n = float(s.replace(",", ""))
    except Exception:
        return s
    for unit, thr in [("T", 1e12), ("G", 1e9), ("M", 1e6), ("K", 1e3)]:
        if abs(n) >= thr:
            return f"{n/thr:.1f}".rstrip("0").rstrip(".") + unit
    return s


def _has_dual_threshold(cfg: ProductConfig) -> bool:
    """判断该产品是否有任意一列配置了双阈值。"""
    return any(len(rule) >= 3 for rule in cfg.qc_rules.values())


# ─────────────────────────────────────────────
# QC_Report（支持单/双阈值）
# ─────────────────────────────────────────────
def write_qc_report(wb, qc_df: pd.DataFrame, fail_dict: dict,
                    risk_dict: dict, cfg: ProductConfig):
    ws = wb.create_sheet("QC_Report")
    headers = list(qc_df.columns)
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    cleandata_col = col_idx.get("CleanData")

    ws.append(headers)
    _style_header(ws, ncols)

    for r_idx, (_, row) in enumerate(qc_df.iterrows(), start=2):
        sample = str(row.get("Sample", ""))
        # 行底色：不合格=红 > 风险=黄 > 其他=无
        if sample in fail_dict:
            row_fill = RED_FILL
        elif sample in risk_dict:
            row_fill = YELLOW_FILL
        else:
            row_fill = None

        row_vals = [
            _bytes_to_human(row.get(c, "")) if c == "CleanData" else row.get(c, "")
            for c in headers
        ]
        ws.append(row_vals)
        _style_row(ws, r_idx, ncols, fill=row_fill)
        if cleandata_col:
            cell = ws.cell(row=r_idx, column=cleandata_col)
            cell.number_format = "@"
            cell.alignment = LEFT

    # 条件格式（按阈值类型分支）
    nrows = len(qc_df) + 1
    for col_name, rule in cfg.qc_rules.items():
        if col_name not in col_idx:
            continue
        cl  = get_column_letter(col_idx[col_name])
        rng = f"{cl}2:{cl}{nrows}"
        pass_val = rule[1]

        if len(rule) >= 3:
            # 双阈值：绿 / 黄 / 红
            risk_val = rule[2]
            ws.conditional_formatting.add(
                rng, CellIsRule("greaterThanOrEqual", [str(pass_val)], fill=GREEN_FILL))
            ws.conditional_formatting.add(
                rng, FormulaRule(
                    formula=[f"AND({cl}2>={risk_val},{cl}2<{pass_val})"],
                    fill=YELLOW_FILL))
            ws.conditional_formatting.add(
                rng, CellIsRule("lessThan", [str(risk_val)], fill=RED_FILL))
        else:
            # 单阈值：绿 / 红
            ws.conditional_formatting.add(
                rng, CellIsRule("greaterThanOrEqual", [str(pass_val)], fill=GREEN_FILL))
            ws.conditional_formatting.add(
                rng, CellIsRule("lessThan", [str(pass_val)], fill=RED_FILL))

    ws.freeze_panes = ws["A2"]
    _auto_col_width(ws)


# ─────────────────────────────────────────────
# QC_FailItems（不合格 + 风险均输出）
# ─────────────────────────────────────────────
def write_qc_failitem(wb, fail_dict: dict, risk_dict: dict):
    ws = wb.create_sheet("QC_FailItems")
    ws.append(["Sample", "Status", "QC_Message"])
    _style_header(ws, 3)

    r_idx = 2
    has_any = False

    for sample, fails in fail_dict.items():
        msg = f"{sample} {', '.join(fails)} 质控不合格"
        ws.append([sample, "不合格", msg])
        _style_row(ws, r_idx, 3, fill=RED_FILL)
        r_idx += 1
        has_any = True

    for sample, risks in risk_dict.items():
        # 风险样本只在该样本没有不合格时输出（已不合格的不重复）
        if sample in fail_dict:
            continue
        msg = f"{sample} {', '.join(risks)} 质控风险"
        ws.append([sample, "风险", msg])
        _style_row(ws, r_idx, 3, fill=YELLOW_FILL)
        r_idx += 1
        has_any = True

    if not has_any:
        ws.append(["—", "合格", "所有样本均通过质控"])
        _style_row(ws, 2, 3, fill=GREEN_FILL)

    ws.freeze_panes = ws["A2"]
    _auto_col_width(ws)


# ─────────────────────────────────────────────
# 通用 SNVIndel 写出（可复用于三种 sheet）
# ─────────────────────────────────────────────
def _write_snv_sheet(wb, sheet_name: str, df: pd.DataFrame,
                     review_cols: list, cfg: ProductConfig,
                     extra_flags: list[str] | None = None):
    """
    通用 SNV 写出。
    extra_flags: 除公共统计列外，额外需要输出的标记列（如 LowFreq_Flag）。
    """
    if df is None or df.empty:
        return

    ws = wb.create_sheet(sheet_name)
    common_extra = ["DisCard_Count", "SNVIndel_Count", "IsFakePositive", "LowAltDepth_Flag"]
    all_extra    = common_extra + (extra_flags or [])
    avail_review = [c for c in review_cols if c in df.columns]
    avail_extra  = [c for c in all_extra if c in df.columns]
    all_cols     = avail_review + avail_extra
    out_df       = df[all_cols].copy()
    ncols        = len(all_cols)
    col_idx      = {h: i + 1 for i, h in enumerate(all_cols)}

    ws.append(all_cols)
    _style_header(ws, ncols)

    for r_idx, (_, row) in enumerate(out_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in all_cols])

        low_flag  = str(row.get("LowAltDepth_Flag", ""))
        low_freq  = str(row.get("LowFreq_Flag", ""))
        fake_pos  = str(row.get("IsFakePositive", ""))
        discard   = to_num(row.get("DisCard_Count", 0)) or 0

        if low_flag:
            fill = ORANGE_FILL
        elif low_freq:
            fill = YELLOW_FILL
        elif fake_pos == "是":
            fill = YELLOW_FILL
        elif discard > 0:
            fill = BLUE_FILL
        else:
            fill = None
        _style_row(ws, r_idx, ncols, fill=fill)

    nrows = len(out_df) + 1
    # AltDepth 条件格式
    if "AltDepth" in col_idx:
        cl = get_column_letter(col_idx["AltDepth"])
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{nrows}",
            FormulaRule(formula=[f"${cl}2<={cfg.low_altdepth_threshold}"],
                        fill=ORANGE_FILL),
        )
    # LowFreq 条件格式
    if "Freq" in col_idx and "LowFreq_Flag" in col_idx:
        cl = get_column_letter(col_idx["Freq"])
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{nrows}",
            FormulaRule(formula=[f"${cl}2<{cfg.hot_somatic_low_freq}"],
                        fill=YELLOW_FILL),
        )
    # DisCard_Count 蓝色加粗
    if "DisCard_Count" in col_idx:
        cl = get_column_letter(col_idx["DisCard_Count"])
        ws.conditional_formatting.add(
            f"{cl}2:{cl}{nrows}",
            CellIsRule("greaterThan", ["0"],
                       font=Font(bold=True, color="0070C0", name="Arial", size=10)),
        )

    ws.freeze_panes = ws["A2"]
    _auto_col_width(ws)
    ws.row_dimensions[1].height = 20


# ─────────────────────────────────────────────
# 标准模式：SNVIndel_Review（OncoPro）
# ─────────────────────────────────────────────
def write_snvindel_review(wb, snv_df: pd.DataFrame, cfg: ProductConfig):
    _write_snv_sheet(wb, "SNVIndel_Review", snv_df, cfg.snv_review_cols, cfg)


# ─────────────────────────────────────────────
# 三表模式（NewProduct）
# ─────────────────────────────────────────────
def write_hot_somatic_review(wb, df: pd.DataFrame, cfg: ProductConfig):
    _write_snv_sheet(wb, "HotSomatic_Review", df,
                     cfg.hot_somatic_review_cols, cfg,
                     extra_flags=["LowFreq_Flag"])


def write_somatic_review(wb, df: pd.DataFrame, cfg: ProductConfig):
    _write_snv_sheet(wb, "Somatic_Review", df, cfg.somatic_review_cols, cfg)


def write_discard_review(wb, df: pd.DataFrame, cfg: ProductConfig):
    """SNVIndelDiscard 二次筛选结果，无公共统计列，直接输出原始列。"""
    if df is None or df.empty:
        return
    ws = wb.create_sheet("Discard_Review")
    avail = [c for c in cfg.discard_review_cols if c in df.columns]
    if not avail:
        avail = list(df.columns)
    ncols = len(avail)
    ws.append(avail)
    _style_header(ws, ncols)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in avail])
        _style_row(ws, r_idx, ncols)
    ws.freeze_panes = ws["A2"]
    _auto_col_width(ws)


# ─────────────────────────────────────────────
# CNV（标记 Confidence=Low）
# ─────────────────────────────────────────────
def write_cnv(wb, cnv_df: pd.DataFrame):
    if cnv_df is None or cnv_df.empty:
        return
    ws      = wb.create_sheet("CNV")
    headers = list(cnv_df.columns)
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    ws.append(headers); _style_header(ws, ncols)
    for r_idx, (_, row) in enumerate(cnv_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        is_low = str(row.get("Confidence", "")).strip().lower() == "low"
        _style_row(ws, r_idx, ncols, fill=ORANGE_FILL if is_low else None)
    if "Confidence" in col_idx:
        cl = get_column_letter(col_idx["Confidence"])
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{len(cnv_df) + 1}",
            FormulaRule(formula=[f'${cl}2="Low"'], fill=ORANGE_FILL),
        )
    ws.freeze_panes = ws["A2"]; _auto_col_width(ws)


# ─────────────────────────────────────────────
# AmpliconStat 透视表
# ─────────────────────────────────────────────
def write_amplicon_pivot(wb, amplicon_df: pd.DataFrame):
    if amplicon_df is None or amplicon_df.empty:
        return
    ws = wb.create_sheet("AmpliconStat")
    required = {"Sample", "Amplicon", "RoT"}
    missing  = required - set(amplicon_df.columns)
    if missing:
        print(f"  ⚠ AmpliconStat 缺少列 {missing}，按原始数据输出")
        headers = list(amplicon_df.columns)
        ws.append(headers); _style_header(ws, len(headers))
        for r_idx, (_, row) in enumerate(amplicon_df.iterrows(), start=2):
            ws.append([row.get(c, "") for c in headers])
            _style_row(ws, r_idx, len(headers))
        ws.freeze_panes = ws["A2"]; _auto_col_width(ws)
        return
    df = amplicon_df.copy()
    df["RoT"] = pd.to_numeric(df["RoT"], errors="coerce")
    pivot = df.pivot_table(index="Amplicon", columns="Sample",
                           values="RoT", aggfunc="mean").reset_index()
    pivot.columns.name = None
    samples = [c for c in pivot.columns if c != "Amplicon"]
    headers = ["Amplicon"] + samples
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    ws.append(headers); _style_header(ws, ncols)
    for r_idx, (_, row) in enumerate(pivot.iterrows(), start=2):
        ws.append([row.get(h, "") for h in headers])
        _style_row(ws, r_idx, ncols)
        for s in samples:
            val = row.get(s)
            cell = ws.cell(row=r_idx, column=col_idx[s])
            if pd.notna(val):
                cell.value = round(float(val), 4)
                cell.number_format = "0.0000"
                cell.alignment = CENTER
        ws.cell(row=r_idx, column=col_idx["Amplicon"]).alignment = LEFT
    ws.freeze_panes = ws["B2"]; _auto_col_width(ws)


# ─────────────────────────────────────────────
# 通用透传
# ─────────────────────────────────────────────
def write_passthrough(wb, sheet_name: str, df: pd.DataFrame):
    if df is None or df.empty:
        return
    ws      = wb.create_sheet(sheet_name)
    headers = list(df.columns)
    ncols   = len(headers)
    ws.append(headers); _style_header(ws, ncols)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        _style_row(ws, r_idx, ncols)
    ws.freeze_panes = ws["A2"]; _auto_col_width(ws)


# ─────────────────────────────────────────────
# HD_pass（含 HD_Flag 标记，Yes 行橙色）
# ─────────────────────────────────────────────
def write_hd_pass(wb, hd_df: pd.DataFrame):
    if hd_df is None or hd_df.empty:
        return
    ws      = wb.create_sheet("HD_pass")
    headers = list(hd_df.columns)
    ncols   = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    ws.append(headers)
    _style_header(ws, ncols)

    for r_idx, (_, row) in enumerate(hd_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in headers])
        fill = ORANGE_FILL if str(row.get("HD_Flag", "")).strip() == "Yes" else None
        _style_row(ws, r_idx, ncols, fill=fill)

    # 条件格式：HD_Flag = "Yes" → 橙色（Excel 中动态生效）
    if "HD_Flag" in col_idx:
        cl = get_column_letter(col_idx["HD_Flag"])
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(ncols)}{len(hd_df) + 1}",
            FormulaRule(formula=[f'${cl}2="Yes"'], fill=ORANGE_FILL),
        )

    ws.freeze_panes = ws["A2"]
    _auto_col_width(ws)
